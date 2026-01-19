"""
Reports API views - attendance, TIL, leave reports with export functionality
"""
import csv
import io
from datetime import datetime, timedelta
from decimal import Decimal

from django.http import HttpResponse
from django.db.models import Sum, Avg, Count, Q, F
from django.db.models.functions import TruncDate, TruncWeek, TruncMonth
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from attendance.models import (
    DailySummary, EmployeeProfile, Department, TILRecord, TILBalance,
    LeaveRecord, ShiftAssignment
)
from .permissions import IsEmployee, IsManager, IsHRAdmin


def get_date_range(request):
    """Parse date range from request params"""
    start_date_str = request.query_params.get('start_date')
    end_date_str = request.query_params.get('end_date')
    preset = request.query_params.get('preset')

    today = timezone.now().date()

    if preset:
        if preset == 'this_week':
            start_date = today - timedelta(days=today.weekday())
            end_date = today
        elif preset == 'last_week':
            start_date = today - timedelta(days=today.weekday() + 7)
            end_date = start_date + timedelta(days=6)
        elif preset == 'this_month':
            start_date = today.replace(day=1)
            end_date = today
        elif preset == 'last_month':
            first_of_month = today.replace(day=1)
            end_date = first_of_month - timedelta(days=1)
            start_date = end_date.replace(day=1)
        elif preset == 'last_quarter':
            quarter = (today.month - 1) // 3
            if quarter == 0:
                start_date = datetime(today.year - 1, 10, 1).date()
                end_date = datetime(today.year - 1, 12, 31).date()
            else:
                start_month = (quarter - 1) * 3 + 1
                start_date = datetime(today.year, start_month, 1).date()
                end_month = quarter * 3
                if end_month == 12:
                    end_date = datetime(today.year, 12, 31).date()
                else:
                    end_date = datetime(today.year, end_month + 1, 1).date() - timedelta(days=1)
        elif preset == 'this_year':
            start_date = datetime(today.year, 1, 1).date()
            end_date = today
        else:
            # Default to last 30 days
            start_date = today - timedelta(days=30)
            end_date = today
    else:
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            start_date = today - timedelta(days=30)

        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = today

    return start_date, end_date


def filter_by_role(queryset, employee_profile, employee_id_field='employee_id'):
    """Filter queryset based on user role"""
    if employee_profile.role == 'HR_ADMIN':
        return queryset

    if employee_profile.role == 'MANAGER':
        team_ids = list(employee_profile.team_members.values_list('employee_id', flat=True))
        team_ids.append(employee_profile.employee_id)
        return queryset.filter(**{f'{employee_id_field}__in': team_ids})

    return queryset.filter(**{employee_id_field: employee_profile.employee_id})


# =============================================================================
# Attendance Reports
# =============================================================================

@api_view(['GET'])
@permission_classes([IsEmployee])
def attendance_report_view(request):
    """
    Get attendance report with filtering and aggregations

    Query params:
    - start_date, end_date: Date range (YYYY-MM-DD)
    - preset: this_week, last_week, this_month, last_month, last_quarter, this_year
    - department: Department ID filter
    - employee: Employee ID filter
    """
    employee_profile = request.user.employee_profile
    start_date, end_date = get_date_range(request)

    # Base queryset
    queryset = DailySummary.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    )

    # Role-based filtering
    queryset = filter_by_role(queryset, employee_profile)

    # Department filter (managers/HR only)
    department_id = request.query_params.get('department')
    if department_id and employee_profile.is_manager_or_above():
        dept_employee_ids = EmployeeProfile.objects.filter(
            department_id=department_id
        ).values_list('employee_id', flat=True)
        queryset = queryset.filter(employee_id__in=dept_employee_ids)

    # Employee filter
    employee_filter = request.query_params.get('employee')
    if employee_filter:
        queryset = queryset.filter(employee_id=employee_filter)

    # Aggregations
    summary = queryset.aggregate(
        total_records=Count('id'),
        total_hours=Sum('final_hours'),
        avg_hours_per_day=Avg('final_hours'),
        total_late_clockouts=Count('id', filter=Q(current_status='IN'))
    )

    # Daily breakdown
    daily_data = queryset.values('date').annotate(
        total_hours=Sum('final_hours'),
        employees_count=Count('employee_id', distinct=True),
        avg_hours=Avg('final_hours')
    ).order_by('date')

    # Employee breakdown (for managers/HR)
    employee_data = []
    if employee_profile.is_manager_or_above():
        employee_data = list(queryset.values('employee_id', 'employee_name').annotate(
            total_hours=Sum('final_hours'),
            days_worked=Count('id'),
            avg_hours=Avg('final_hours')
        ).order_by('-total_hours'))

    # Detailed records (paginated)
    page = int(request.query_params.get('page', 1))
    page_size = int(request.query_params.get('page_size', 50))
    offset = (page - 1) * page_size

    records = list(queryset.order_by('-date', 'employee_name')[offset:offset + page_size].values(
        'id', 'date', 'employee_id', 'employee_name',
        'first_clock_in', 'last_clock_out', 'final_hours',
        'current_status', 'tap_count'
    ))

    return Response({
        'date_range': {
            'start_date': str(start_date),
            'end_date': str(end_date)
        },
        'summary': {
            'total_records': summary['total_records'] or 0,
            'total_hours': float(summary['total_hours'] or 0),
            'avg_hours_per_day': float(summary['avg_hours_per_day'] or 0),
            'still_clocked_in': summary['total_late_clockouts'] or 0
        },
        'daily_breakdown': list(daily_data),
        'employee_breakdown': employee_data,
        'records': records,
        'pagination': {
            'page': page,
            'page_size': page_size,
            'total_records': queryset.count()
        }
    })


@api_view(['GET'])
@permission_classes([IsEmployee])
def attendance_export_view(request):
    """
    Export attendance report as CSV, Excel, or PDF

    Query params:
    - format: csv, excel, pdf (default: csv)
    - ... (same filters as attendance_report_view)
    """
    employee_profile = request.user.employee_profile
    start_date, end_date = get_date_range(request)
    export_format = request.query_params.get('format', 'csv')

    # Base queryset
    queryset = DailySummary.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    )

    # Role-based filtering
    queryset = filter_by_role(queryset, employee_profile)

    # Department filter
    department_id = request.query_params.get('department')
    if department_id and employee_profile.is_manager_or_above():
        dept_employee_ids = EmployeeProfile.objects.filter(
            department_id=department_id
        ).values_list('employee_id', flat=True)
        queryset = queryset.filter(employee_id__in=dept_employee_ids)

    # Employee filter
    employee_filter = request.query_params.get('employee')
    if employee_filter:
        queryset = queryset.filter(employee_id=employee_filter)

    queryset = queryset.order_by('-date', 'employee_name')

    if export_format == 'csv':
        return export_attendance_csv(queryset, start_date, end_date)
    elif export_format == 'excel':
        return export_attendance_excel(queryset, start_date, end_date)
    elif export_format == 'pdf':
        return export_attendance_pdf(queryset, start_date, end_date)
    else:
        return Response({'error': 'Invalid format'}, status=status.HTTP_400_BAD_REQUEST)


def export_attendance_csv(queryset, start_date, end_date):
    """Export attendance as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="attendance_{start_date}_{end_date}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Date', 'Employee ID', 'Employee Name', 'Clock In', 'Clock Out',
                     'Hours Worked', 'Status', 'Taps'])

    for record in queryset:
        writer.writerow([
            record.date,
            record.employee_id,
            record.employee_name,
            record.first_clock_in or '',
            record.last_clock_out or '',
            record.final_hours,
            record.current_status,
            record.tap_count
        ])

    return response


def export_attendance_excel(queryset, start_date, end_date):
    """Export attendance as Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return Response(
            {'error': 'Excel export not available. openpyxl not installed.'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f'Attendance Report: {start_date} to {end_date}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Date', 'Employee ID', 'Employee Name', 'Clock In', 'Clock Out',
               'Hours Worked', 'Status', 'Taps']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Data
    for row_num, record in enumerate(queryset, 4):
        ws.cell(row=row_num, column=1, value=str(record.date)).border = border
        ws.cell(row=row_num, column=2, value=record.employee_id).border = border
        ws.cell(row=row_num, column=3, value=record.employee_name).border = border
        ws.cell(row=row_num, column=4, value=str(record.first_clock_in) if record.first_clock_in else '').border = border
        ws.cell(row=row_num, column=5, value=str(record.last_clock_out) if record.last_clock_out else '').border = border
        ws.cell(row=row_num, column=6, value=float(record.final_hours)).border = border
        ws.cell(row=row_num, column=7, value=record.current_status).border = border
        ws.cell(row=row_num, column=8, value=record.tap_count).border = border

    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 8

    # Save to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="attendance_{start_date}_{end_date}.xlsx"'
    return response


def export_attendance_pdf(queryset, start_date, end_date):
    """Export attendance as PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                           rightMargin=0.5*inch, leftMargin=0.5*inch,
                           topMargin=0.5*inch, bottomMargin=0.5*inch)

    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=20,
        alignment=1  # Center
    )
    elements.append(Paragraph(f'Attendance Report: {start_date} to {end_date}', title_style))
    elements.append(Spacer(1, 12))

    # Table data
    data = [['Date', 'Employee ID', 'Employee Name', 'Clock In', 'Clock Out', 'Hours', 'Status']]

    for record in queryset[:500]:  # Limit to 500 for PDF
        data.append([
            str(record.date),
            record.employee_id,
            record.employee_name[:25],  # Truncate long names
            str(record.first_clock_in)[:5] if record.first_clock_in else '-',
            str(record.last_clock_out)[:5] if record.last_clock_out else '-',
            f'{record.final_hours:.1f}',
            record.current_status
        ])

    # Create table
    table = Table(data, colWidths=[1*inch, 1*inch, 2.5*inch, 0.8*inch, 0.8*inch, 0.7*inch, 0.7*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="attendance_{start_date}_{end_date}.pdf"'
    return response


# =============================================================================
# TIL Reports
# =============================================================================

@api_view(['GET'])
@permission_classes([IsEmployee])
def til_report_view(request):
    """
    Get TIL report with balances, earned, and used breakdown
    """
    employee_profile = request.user.employee_profile
    start_date, end_date = get_date_range(request)

    # TIL Records queryset
    queryset = TILRecord.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    )

    # Role-based filtering using employee FK
    if employee_profile.role == 'HR_ADMIN':
        pass  # No filter
    elif employee_profile.role == 'MANAGER':
        team_ids = list(employee_profile.team_members.values_list('id', flat=True))
        team_ids.append(employee_profile.id)
        queryset = queryset.filter(employee_id__in=team_ids)
    else:
        queryset = queryset.filter(employee=employee_profile)

    # Department filter
    department_id = request.query_params.get('department')
    if department_id and employee_profile.is_manager_or_above():
        queryset = queryset.filter(employee__department_id=department_id)

    # Summary aggregations
    summary = queryset.filter(status='APPROVED').aggregate(
        total_earned_early=Sum('hours', filter=Q(til_type='EARNED_EARLY')),
        total_earned_ot=Sum('hours', filter=Q(til_type='EARNED_OT')),
        total_used=Sum('hours', filter=Q(til_type='USED')),
        total_adjusted=Sum('hours', filter=Q(til_type='ADJUSTED'))
    )

    pending_count = queryset.filter(status='PENDING').count()

    # By type breakdown
    by_type = list(queryset.filter(status='APPROVED').values('til_type').annotate(
        total_hours=Sum('hours'),
        count=Count('id')
    ))

    # By employee breakdown (managers/HR only)
    by_employee = []
    if employee_profile.is_manager_or_above():
        by_employee = list(queryset.filter(status='APPROVED').values(
            'employee__employee_id', 'employee__employee_name'
        ).annotate(
            total_earned=Sum('hours', filter=Q(til_type__in=['EARNED_EARLY', 'EARNED_OT'])),
            total_used=Sum('hours', filter=Q(til_type='USED'))
        ).order_by('-total_earned'))

    # TIL Balances
    balances = []
    if employee_profile.is_manager_or_above():
        balance_qs = TILBalance.objects.select_related('employee')
        if department_id:
            balance_qs = balance_qs.filter(employee__department_id=department_id)
        if employee_profile.role == 'MANAGER':
            team_ids = list(employee_profile.team_members.values_list('id', flat=True))
            team_ids.append(employee_profile.id)
            balance_qs = balance_qs.filter(employee_id__in=team_ids)

        balances = list(balance_qs.values(
            'employee__employee_id', 'employee__employee_name',
            'total_earned', 'total_used', 'current_balance'
        ).order_by('-current_balance'))
    else:
        # Employee sees only their own balance
        try:
            bal = TILBalance.objects.get(employee=employee_profile)
            balances = [{
                'employee__employee_id': employee_profile.employee_id,
                'employee__employee_name': employee_profile.employee_name,
                'total_earned': bal.total_earned,
                'total_used': bal.total_used,
                'current_balance': bal.current_balance
            }]
        except TILBalance.DoesNotExist:
            pass

    # Recent records
    recent_records = list(queryset.order_by('-date', '-created_at')[:20].values(
        'id', 'employee__employee_name', 'til_type', 'status', 'hours', 'date', 'reason'
    ))

    return Response({
        'date_range': {
            'start_date': str(start_date),
            'end_date': str(end_date)
        },
        'summary': {
            'total_earned_early': float(summary['total_earned_early'] or 0),
            'total_earned_overtime': float(summary['total_earned_ot'] or 0),
            'total_used': float(abs(summary['total_used'] or 0)),
            'total_adjusted': float(summary['total_adjusted'] or 0),
            'pending_approvals': pending_count
        },
        'by_type': by_type,
        'by_employee': by_employee,
        'balances': balances,
        'recent_records': recent_records
    })


@api_view(['GET'])
@permission_classes([IsEmployee])
def til_export_view(request):
    """Export TIL report"""
    employee_profile = request.user.employee_profile
    start_date, end_date = get_date_range(request)
    export_format = request.query_params.get('format', 'csv')

    queryset = TILRecord.objects.filter(
        date__gte=start_date,
        date__lte=end_date,
        status='APPROVED'
    ).select_related('employee')

    # Role-based filtering
    if employee_profile.role == 'MANAGER':
        team_ids = list(employee_profile.team_members.values_list('id', flat=True))
        team_ids.append(employee_profile.id)
        queryset = queryset.filter(employee_id__in=team_ids)
    elif employee_profile.role == 'EMPLOYEE':
        queryset = queryset.filter(employee=employee_profile)

    queryset = queryset.order_by('-date')

    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="til_report_{start_date}_{end_date}.csv"'

        writer = csv.writer(response)
        writer.writerow(['Date', 'Employee ID', 'Employee Name', 'Type', 'Hours', 'Reason'])

        for record in queryset:
            writer.writerow([
                record.date,
                record.employee.employee_id,
                record.employee.employee_name,
                record.get_til_type_display(),
                record.hours,
                record.reason
            ])

        return response

    return Response({'error': 'Only CSV format supported for TIL export'}, status=status.HTTP_400_BAD_REQUEST)


# =============================================================================
# Leave Reports
# =============================================================================

@api_view(['GET'])
@permission_classes([IsEmployee])
def leave_report_view(request):
    """
    Get leave report with breakdown by type and status
    """
    employee_profile = request.user.employee_profile
    start_date, end_date = get_date_range(request)

    queryset = LeaveRecord.objects.filter(
        start_date__lte=end_date,
        end_date__gte=start_date
    )

    # Role-based filtering
    if employee_profile.role == 'HR_ADMIN':
        pass
    elif employee_profile.role == 'MANAGER':
        team_ids = list(employee_profile.team_members.values_list('id', flat=True))
        team_ids.append(employee_profile.id)
        queryset = queryset.filter(employee_profile_id__in=team_ids)
    else:
        queryset = queryset.filter(employee_profile=employee_profile)

    # Department filter
    department_id = request.query_params.get('department')
    if department_id and employee_profile.is_manager_or_above():
        queryset = queryset.filter(employee_profile__department_id=department_id)

    # Summary
    summary = queryset.aggregate(
        total_requests=Count('id'),
        approved_count=Count('id', filter=Q(status='APPROVED')),
        pending_count=Count('id', filter=Q(status='PENDING')),
        rejected_count=Count('id', filter=Q(status='REJECTED')),
        total_days_approved=Sum('total_days', filter=Q(status='APPROVED'))
    )

    # By type
    by_type = list(queryset.filter(status='APPROVED').values('leave_type').annotate(
        count=Count('id'),
        total_days=Sum('total_days'),
        total_hours=Sum('total_hours')
    ))

    # By status
    by_status = list(queryset.values('status').annotate(
        count=Count('id')
    ))

    # By employee (managers/HR)
    by_employee = []
    if employee_profile.is_manager_or_above():
        by_employee = list(queryset.filter(status='APPROVED').values(
            'employee_id', 'employee_name'
        ).annotate(
            total_days=Sum('total_days'),
            request_count=Count('id')
        ).order_by('-total_days'))

    # Recent leaves
    recent = list(queryset.order_by('-created_at')[:20].values(
        'id', 'employee_name', 'leave_type', 'status',
        'start_date', 'end_date', 'total_days'
    ))

    return Response({
        'date_range': {
            'start_date': str(start_date),
            'end_date': str(end_date)
        },
        'summary': {
            'total_requests': summary['total_requests'] or 0,
            'approved': summary['approved_count'] or 0,
            'pending': summary['pending_count'] or 0,
            'rejected': summary['rejected_count'] or 0,
            'total_days_approved': summary['total_days_approved'] or 0
        },
        'by_type': by_type,
        'by_status': by_status,
        'by_employee': by_employee,
        'recent_leaves': recent
    })


@api_view(['GET'])
@permission_classes([IsEmployee])
def leave_export_view(request):
    """Export leave report"""
    employee_profile = request.user.employee_profile
    start_date, end_date = get_date_range(request)
    export_format = request.query_params.get('format', 'csv')

    queryset = LeaveRecord.objects.filter(
        start_date__lte=end_date,
        end_date__gte=start_date
    ).select_related('employee_profile')

    # Role-based filtering
    if employee_profile.role == 'MANAGER':
        team_ids = list(employee_profile.team_members.values_list('id', flat=True))
        team_ids.append(employee_profile.id)
        queryset = queryset.filter(employee_profile_id__in=team_ids)
    elif employee_profile.role == 'EMPLOYEE':
        queryset = queryset.filter(employee_profile=employee_profile)

    queryset = queryset.order_by('-start_date')

    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="leave_report_{start_date}_{end_date}.csv"'

        writer = csv.writer(response)
        writer.writerow(['Employee ID', 'Employee Name', 'Leave Type', 'Start Date',
                        'End Date', 'Days', 'Status', 'Reason'])

        for record in queryset:
            writer.writerow([
                record.employee_id,
                record.employee_name,
                record.get_leave_type_display(),
                record.start_date,
                record.end_date,
                record.total_days,
                record.get_status_display(),
                record.reason
            ])

        return response

    return Response({'error': 'Only CSV format supported for leave export'}, status=status.HTTP_400_BAD_REQUEST)


# =============================================================================
# Team Reports (Manager Dashboard)
# =============================================================================

@api_view(['GET'])
@permission_classes([IsManager])
def team_report_view(request):
    """
    Get team report for managers - overview of team attendance, TIL, and leave
    """
    employee_profile = request.user.employee_profile
    today = timezone.now().date()

    # Get team members
    if employee_profile.role == 'HR_ADMIN':
        team_members = EmployeeProfile.objects.filter(is_active=True)
    else:
        team_members = employee_profile.team_members.filter(is_active=True)

    team_ids = list(team_members.values_list('employee_id', flat=True))
    team_profile_ids = list(team_members.values_list('id', flat=True))

    # Today's attendance
    today_attendance = DailySummary.objects.filter(
        date=today,
        employee_id__in=team_ids
    )

    clocked_in = today_attendance.filter(current_status='IN').count()
    clocked_out = today_attendance.filter(current_status='OUT').count()
    not_clocked = len(team_ids) - today_attendance.count()

    # This week's hours
    week_start = today - timedelta(days=today.weekday())
    week_attendance = DailySummary.objects.filter(
        date__gte=week_start,
        date__lte=today,
        employee_id__in=team_ids
    ).aggregate(
        total_hours=Sum('final_hours'),
        avg_hours=Avg('final_hours')
    )

    # Pending approvals
    pending_til = TILRecord.objects.filter(
        employee_id__in=team_profile_ids,
        status='PENDING'
    ).count()

    pending_leave = LeaveRecord.objects.filter(
        employee_profile_id__in=team_profile_ids,
        status='PENDING'
    ).count()

    # TIL balances
    til_balances = list(TILBalance.objects.filter(
        employee_id__in=team_profile_ids
    ).select_related('employee').values(
        'employee__employee_id', 'employee__employee_name', 'current_balance'
    ).order_by('-current_balance')[:10])

    # Upcoming leave
    upcoming_leave = list(LeaveRecord.objects.filter(
        employee_profile_id__in=team_profile_ids,
        status='APPROVED',
        start_date__gte=today,
        start_date__lte=today + timedelta(days=14)
    ).values(
        'employee_name', 'leave_type', 'start_date', 'end_date', 'total_days'
    ).order_by('start_date')[:10])

    # Team members list with today's status
    team_list = []
    for member in team_members[:20]:
        try:
            summary = DailySummary.objects.get(date=today, employee_id=member.employee_id)
            status_today = summary.current_status
            hours_today = float(summary.final_hours)
        except DailySummary.DoesNotExist:
            status_today = 'NOT_CLOCKED'
            hours_today = 0

        team_list.append({
            'employee_id': member.employee_id,
            'employee_name': member.employee_name,
            'department': member.department.name if member.department else None,
            'status_today': status_today,
            'hours_today': hours_today
        })

    return Response({
        'date': str(today),
        'team_size': len(team_ids),
        'today': {
            'clocked_in': clocked_in,
            'clocked_out': clocked_out,
            'not_clocked_in': not_clocked
        },
        'this_week': {
            'total_hours': float(week_attendance['total_hours'] or 0),
            'avg_hours_per_day': float(week_attendance['avg_hours'] or 0)
        },
        'pending_approvals': {
            'til': pending_til,
            'leave': pending_leave,
            'total': pending_til + pending_leave
        },
        'til_balances': til_balances,
        'upcoming_leave': upcoming_leave,
        'team_members': team_list
    })


# =============================================================================
# Analytics / Charts Data
# =============================================================================

@api_view(['GET'])
@permission_classes([IsEmployee])
def attendance_trends_view(request):
    """
    Get attendance trends for charts - hours worked over time
    """
    employee_profile = request.user.employee_profile
    start_date, end_date = get_date_range(request)
    group_by = request.query_params.get('group_by', 'day')  # day, week, month

    queryset = DailySummary.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    )

    # Role-based filtering
    queryset = filter_by_role(queryset, employee_profile)

    # Department filter
    department_id = request.query_params.get('department')
    if department_id and employee_profile.is_manager_or_above():
        dept_employee_ids = EmployeeProfile.objects.filter(
            department_id=department_id
        ).values_list('employee_id', flat=True)
        queryset = queryset.filter(employee_id__in=dept_employee_ids)

    # Group by period
    if group_by == 'week':
        data = queryset.annotate(period=TruncWeek('date')).values('period').annotate(
            total_hours=Sum('final_hours'),
            avg_hours=Avg('final_hours'),
            records=Count('id')
        ).order_by('period')
    elif group_by == 'month':
        data = queryset.annotate(period=TruncMonth('date')).values('period').annotate(
            total_hours=Sum('final_hours'),
            avg_hours=Avg('final_hours'),
            records=Count('id')
        ).order_by('period')
    else:  # day
        data = queryset.values('date').annotate(
            total_hours=Sum('final_hours'),
            avg_hours=Avg('final_hours'),
            records=Count('id')
        ).order_by('date')

    # Format for charts
    chart_data = []
    for item in data:
        period = item.get('period') or item.get('date')
        chart_data.append({
            'date': str(period),
            'total_hours': float(item['total_hours'] or 0),
            'avg_hours': float(item['avg_hours'] or 0),
            'records': item['records']
        })

    return Response({
        'date_range': {
            'start_date': str(start_date),
            'end_date': str(end_date)
        },
        'group_by': group_by,
        'data': chart_data
    })


@api_view(['GET'])
@permission_classes([IsManager])
def department_comparison_view(request):
    """
    Get department comparison data for charts (managers/HR only)
    """
    employee_profile = request.user.employee_profile
    start_date, end_date = get_date_range(request)

    # Get all departments
    departments = Department.objects.filter(is_active=True)

    results = []
    for dept in departments:
        dept_employee_ids = EmployeeProfile.objects.filter(
            department=dept
        ).values_list('employee_id', flat=True)

        stats = DailySummary.objects.filter(
            date__gte=start_date,
            date__lte=end_date,
            employee_id__in=dept_employee_ids
        ).aggregate(
            total_hours=Sum('final_hours'),
            avg_hours=Avg('final_hours'),
            total_records=Count('id'),
            employees_count=Count('employee_id', distinct=True)
        )

        results.append({
            'department': dept.name,
            'code': dept.code,
            'total_hours': float(stats['total_hours'] or 0),
            'avg_hours': float(stats['avg_hours'] or 0),
            'total_records': stats['total_records'] or 0,
            'employees': stats['employees_count'] or 0
        })

    return Response({
        'date_range': {
            'start_date': str(start_date),
            'end_date': str(end_date)
        },
        'departments': results
    })
